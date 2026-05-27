"""
app.py
DART 재무 데이터 분석기 - Streamlit 메인 UI

여러 기업명과 사업연도를 입력하면 DART API로 재무 데이터를 자동 추출하고
회계 항등식·교차 검증까지 수행하여 결과를 표·상세뷰·엑셀로 출력한다.
"""

import io
import json
import re
from itertools import product
from typing import Optional

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from config import FINANCIAL_ITEMS, get_dart_api_key, get_gemini_api_key
from dart_api.corp_search import download_corp_codes, get_company_info_batch, search_corp
from pipeline import analyze_one as _analyze_one
from utils.analysis_logger import format_amount

# ── 페이지 설정 ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DART 재무 데이터 분석기",
    page_icon="📊",
    layout="wide",
)

# ── localStorage → session_state 복원 ────────────────────────────────────────
# 첫 로드 시 브라우저 localStorage에 저장된 API 키를 query_params 경유로 복원한다.
_LS_KEYS = ("_dk", "_gk")

if any(k in st.query_params for k in _LS_KEYS):
    if "_dk" in st.query_params:
        st.session_state.dart_api_key = st.query_params["_dk"]
    if "_gk" in st.query_params:
        st.session_state.gemini_api_key = st.query_params["_gk"]
    for k in _LS_KEYS:
        st.query_params.pop(k, None)
    st.rerun()

if "dart_api_key" not in st.session_state:
    components.html("""
    <script>
    const dk = localStorage.getItem('dart_api_key') || '';
    const gk = localStorage.getItem('gemini_api_key') || '';
    if (dk || gk) {
        const url = new URL(window.parent.location);
        if (dk) url.searchParams.set('_dk', dk);
        if (gk) url.searchParams.set('_gk', gk);
        window.parent.location.replace(url.toString());
    }
    </script>
    """, height=0)

# ── 세션 상태 초기화 ──────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results: list[dict] = []
if "analysis_logs" not in st.session_state:
    st.session_state.analysis_logs: list[str] = []

# ── 표시 컬럼 정의 ────────────────────────────────────────────────────────────
# 요약 테이블에 표시할 재무 항목 (순서 유지)
_DISPLAY_ITEMS = [
    "총자산", "총부채", "이익잉여금",
    "감가상각비", "사용권자산상각비", "무형자산상각비", "EBITDA",
    "매출액", "매출총이익", "영업이익", "당기순이익",
]

# EBITDA 계산은 financial.ebitda를 단일 소스로 공유한다.
from financial.ebitda import (
    compute_ebitda as _compute_ebitda,
    EBITDA_OPERATING_INCOME as _EBITDA_OP,
    EBITDA_ADDBACKS as _EBITDA_ADDBACKS,
)



# ── 금액 포맷 ─────────────────────────────────────────────────────────────────

def _fmt_억(val: Optional[float]) -> str:
    """원(KRW) 값을 억 단위 문자열로 변환한다. 음수는 괄호 표기."""
    if val is None:
        return "-"
    억 = val / 1e8
    if 억 < 0:
        return f"({abs(억):,.0f})"
    return f"{억:,.0f}"


def _fmt_억_raw(val: Optional[float]) -> Optional[float]:
    """Excel 출력용: 억 단위 float 반환 (반올림 없이 원본 정밀도 유지)."""
    return val / 1e8 if val is not None else None


def _item_value(items: dict, name: str) -> Optional[float]:
    """표시용 항목 값 조회. EBITDA는 원본에 없는 파생 항목이라 직접 계산한다."""
    if name == "EBITDA":
        return _compute_ebitda(items)
    return items.get(name)


def _sanitize_filename_part(text: str) -> str:
    """파일명에 안전한 문자열만 남긴다."""
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", (text or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned.strip("._") or "trace"


def _depreciation_trace_filename(result: dict) -> str:
    """결과 1건에 대한 감가상각 추적 로그 파일명을 생성한다."""
    corp = _sanitize_filename_part(result.get("corp_name", "corp"))
    year = result.get("year", "year")
    path = result.get("path", "-")
    report = _sanitize_filename_part(result.get("report_nm", "report"))
    return f"{corp}_{year}_경로{path}_{report}_감가상각추적.txt"


def _depreciation_trace_text(result: dict) -> str:
    """결과 1건의 감가상각 추적 로그를 TXT 본문으로 직렬화한다."""
    trace_lines = result.get("depreciation_trace") or []
    header = [
        f"기업명: {result.get('corp_name', '-')}",
        f"연도: {result.get('year', '-')}",
        f"corp_code: {result.get('corp_code', '-')}",
        f"경로: {result.get('path', '-')}",
        f"보고서: {result.get('report_nm', '-')}",
        f"재무제표기준: {_fs_label(result)}",
        "",
        "[감가상각 추적 로그]",
    ]
    if trace_lines:
        return "\n".join(header + trace_lines)
    return "\n".join(header + ["로그 없음"])


def _result_validation_status(result: dict) -> Optional[str]:
    """결과의 검증 상태(verified/partial/failed_validation/unverified)를 반환한다.

    구버전 결과(validation_status 없음)는 is_valid로 폴백한다.
    """
    val = result.get("validation")
    if not val:
        return None
    status = val.get("validation_status")
    if status is not None:
        return status
    return "verified" if val.get("is_valid") else "failed_validation"


def _validation_status_label(result: dict, use_icon: bool = True) -> str:
    """결과 딕셔너리의 검증 상태를 짧은 라벨로 반환한다."""
    if result["status"] != "ok":
        return result["error_msg"]
    status = _result_validation_status(result)
    if status is None:
        return "검증정보 없음"

    label, icon = {
        "verified":         ("통과", "✓"),
        "partial":          ("일부검증", "△"),
        "failed_validation": ("검증실패", "✗"),
        "unverified":       ("검증불가", "?"),
    }.get(status, ("검증불가", "?"))

    return f"{icon} {label}" if use_icon else label


def _processing_status_detail(result: dict) -> str:
    """분석/추출 단계 상태를 사람이 읽기 쉬운 설명으로 반환한다."""
    status = result.get("status")
    if status == "ok":
        report_nm = result.get("report_nm") or "-"
        return (
            f"데이터 추출 성공. 사용 보고서: {report_nm}, "
            f"경로 {result.get('path', '-')}, 재무제표 기준 {_fs_label(result)}."
        )

    if status == "no_corp":
        msg = result.get("error_msg") or "기업명을 찾지 못했습니다."
        if "유사:" in msg:
            return (
                f"{msg}. 입력한 이름과 정확히 일치하는 DART 기업명이 없어 자동 선택하지 않았습니다. "
                "동명이인이나 유사 상호가 많을 수 있으니 정식 법인명으로 다시 조회해 주세요."
            )
        return (
            f"{msg}. DART 기업코드 목록에서 정확 일치 기업명을 찾지 못했습니다."
        )

    if status == "ambiguous_corp":
        msg = result.get("error_msg") or "동일 기업명이 여러 개 존재합니다."
        return f"{msg} DART에 같은 이름의 법인이 다수 등록되어 있어 재무데이터를 특정할 수 없습니다."

    if status == "no_report":
        year = result.get("year")
        return (
            f"{year} 사업연도 기준 사업보고서, 연결감사보고서, 감사보고서를 순서대로 찾았지만 "
            "사용 가능한 보고서를 찾지 못했습니다."
        )

    if status == "error":
        msg = result.get("error_msg") or "분석 중 오류가 발생했습니다."
        if "보고서 탐색 오류" in msg:
            return f"{msg}. DART 공시 목록 조회 단계에서 실패했습니다."
        if "재무데이터 추출 오류" in msg:
            return f"{msg}. 보고서는 찾았지만 재무 항목 추출 단계에서 실패했습니다."
        if "document.xml 다운로드 실패" in msg:
            return f"{msg}. 감사보고서 원문(document.xml)을 내려받지 못했습니다."
        return msg

    return result.get("error_msg") or "상세 사유 정보가 없습니다."


def _validation_reason_lines(result: dict, include_ai: bool = True) -> list[str]:
    """결과 딕셔너리의 검증/처리 사유를 상세 문장 리스트로 반환한다."""
    if result["status"] != "ok":
        return [_processing_status_detail(result)]

    val = result.get("validation")
    if not val:
        return ["검증 정보가 없습니다."]

    lines: list[str] = []
    failed_checks = [c for c in val.get("checks", []) if c.get("status") == "failed"]
    skipped_checks = [c for c in val.get("checks", []) if c.get("status") == "skipped_missing_data"]

    if failed_checks:
        severity_map = {"critical": "치명", "warning": "경고", "info": "참고"}
        for check in failed_checks:
            severity = severity_map.get(check.get("severity"), str(check.get("severity", "")).upper())
            line = f"[{severity}] {check.get('rule', '-')}"
            diff = check.get("diff")
            if isinstance(diff, (int, float)):
                line += f" (차이 {diff:,.0f}원)"
            note = check.get("note")
            if note:
                line += f" - {note}"
            lines.append(line)
    else:
        lines.append("실패한 검증은 없습니다.")

    if skipped_checks:
        for check in skipped_checks:
            lines.append(f"[스킵] {check.get('rule', '-')} - {check.get('note', '사유 없음')}")

    # AI 비교 결과 (경로A·경로B 공통)
    ai_comp = result.get("ai_comparison")
    if include_ai and ai_comp:
        source_label = {
            "agreed": "AI 비교: Python·AI 일치",
            "adjudicated": "AI 비교: AI 판정으로 불일치 해소",
            "ai_extract_only": "AI 비교: AI 추출값 채택",
            "python_fallback": "AI 비교: AI 실패로 Python 결과 사용",
        }.get(ai_comp.get("source"), f"AI 비교: {ai_comp.get('source', '비교 수행')}")
        if ai_comp.get("disagreements"):
            source_label += f" (불일치 항목: {', '.join(ai_comp['disagreements'].keys())})"
        if ai_comp.get("error"):
            source_label += f" / AI 오류: {ai_comp['error']}"
        lines.append(source_label)

    return lines


def _format_validation_detail(result: dict, include_ai: bool = True) -> str:
    """결과 딕셔너리의 검증 상세를 한 줄 문자열로 반환한다."""
    return " | ".join(_validation_reason_lines(result, include_ai=include_ai))


# ── 결과 → DataFrame ──────────────────────────────────────────────────────────

def _fs_label(r: dict) -> str:
    """결과 딕셔너리에서 재무제표기준(연결/별도)을 반환한다."""
    if r["path"] == "A":
        return "연결" if r.get("fs_div") == "CFS" else "별도"
    # 경로B: report_type으로 판단
    rt = r.get("report_type", "-")
    if rt == "audit_consol":
        return "연결"
    if rt == "audit_separate":
        return "별도"
    return "-"


def _to_summary_df(results: list[dict]) -> pd.DataFrame:
    """결과 리스트를 요약 표시용 DataFrame으로 변환한다."""
    rows = []
    for r in results:
        items = r.get("items", {})

        row = {
            "기업명":       r["corp_name"],
            "연도":         r["year"],
            "재무제표기준": _fs_label(r),
            "경로":         r["path"],
            "보고서":       r["report_nm"],
        }
        for name in _DISPLAY_ITEMS:
            row[f"{name}(억)"] = _fmt_억(_item_value(items, name))
        row["검증"] = _validation_status_label(r, use_icon=True)
        row["처리상세"] = _processing_status_detail(r)
        row["검증상세"] = _format_validation_detail(r)
        rows.append(row)

    return pd.DataFrame(rows)


# ── 결과 → Excel bytes ────────────────────────────────────────────────────────

def _to_excel_summary_df(results: list[dict]) -> pd.DataFrame:
    """엑셀용 요약 DataFrame: 금액 컬럼을 억 단위 숫자(float)로 반환한다."""
    rows = []
    for r in results:
        items = r.get("items", {})

        row = {
            "기업명":       r["corp_name"],
            "연도":         r["year"],
            "재무제표기준": _fs_label(r),
            "경로":         r["path"],
            "보고서":       r["report_nm"],
        }
        for name in _DISPLAY_ITEMS:
            row[f"{name}(억원)"] = _fmt_억_raw(_item_value(items, name))
        row["검증"] = _validation_status_label(r, use_icon=False)
        row["처리상세"] = _processing_status_detail(r)
        row["검증상세"] = _format_validation_detail(r)
        rows.append(row)

    return pd.DataFrame(rows)


def _apply_number_format(ws, df: pd.DataFrame, fmt: str = "#,##0_);(#,##0);-_)") -> None:
    """워크시트에서 숫자형 컬럼에 엑셀 숫자 서식을 적용한다."""
    from openpyxl.styles import numbers as xl_numbers
    numeric_cols = [i + 1 for i, col in enumerate(df.columns) if pd.api.types.is_float_dtype(df[col]) or pd.api.types.is_integer_dtype(df[col])]
    for col_idx in numeric_cols:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = fmt


def _to_excel_bytes(results: list[dict]) -> bytes:
    """결과 리스트를 Excel bytes로 변환한다.

    시트 구성: 회사별 1시트 (가로=연도, 세로=계정항목) + 원본 시트 1개.
    """
    from openpyxl.styles import Alignment, Font, PatternFill, numbers as xl_numbers

    # 항목 표시 순서: 매출 먼저, 그 다음 BS
    _EXCEL_ITEMS = ["매출액", "매출총이익", "영업이익", "당기순이익",
                    "총자산", "총부채", "이익잉여금",
                    "감가상각비", "사용권자산상각비", "무형자산상각비", "EBITDA"]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── 회사별 시트 ──────────────────────────────────────────────────
        # 결과를 기업별로 그룹핑
        from collections import defaultdict
        corp_results: dict[str, list[dict]] = defaultdict(list)
        for r in results:
            corp_results[r["corp_name"]].append(r)

        for corp_name, corp_data in corp_results.items():
            # 연도 오름차순 정렬
            corp_data.sort(key=lambda x: x["year"])
            years = [r["year"] for r in corp_data]

            # DataFrame 구성: 항목명 | 2022 | 2023 | 2024 ...
            rows = []
            for item_name in _EXCEL_ITEMS:
                row: dict = {"항목": item_name}
                for r in corp_data:
                    yr = r["year"]
                    val = r.get("items", {}).get(item_name)
                    row[str(yr)] = _fmt_억_raw(val)
                rows.append(row)

            df = pd.DataFrame(rows)

            # 시트명은 최대 31자, 특수문자 제거
            sheet_name = corp_name[:28]
            # 중복 시트명 방지
            existing = [s for s in writer.sheets]
            if sheet_name in existing:
                sheet_name = sheet_name[:25] + f"_{len(existing)}"

            df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
            ws = writer.sheets[sheet_name]

            # 헤더 행1: 기업명
            ws.cell(row=1, column=1, value=corp_name).font = Font(bold=True, size=13)

            # 헤더 행2: 각 연도 아래 연결/별도 + 경로 표시
            for col_idx, r in enumerate(corp_data):
                fs = _fs_label(r)
                path = r["path"]
                cell = ws.cell(row=2, column=col_idx + 2,
                               value=f"{fs} (경로{path})")
                cell.font = Font(italic=True, color="666666", size=9)
                cell.alignment = Alignment(horizontal="center")

            # 연도 헤더(행3)는 pandas가 이미 기록 — 가운데 정렬 + 볼드
            for col_idx in range(len(years)):
                cell = ws.cell(row=3, column=col_idx + 2)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # 데이터 영역 숫자 서식 적용
            for row_idx in range(4, 4 + len(_EXCEL_ITEMS)):
                for col_idx in range(2, 2 + len(years)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = "#,##0_);(#,##0);-_)"
                        cell.alignment = Alignment(horizontal="right")

            # EBITDA 행: 하드코딩 값 대신 셀 참조 수식으로 기입
            # (영업이익 + 감가상각비 + 사용권자산상각비 + 무형자산상각비)
            from openpyxl.utils import get_column_letter
            _row_of = lambda name: 4 + _EXCEL_ITEMS.index(name)
            ebitda_row = _row_of("EBITDA")
            op_row = _row_of(_EBITDA_OP)
            addback_rows = [_row_of(k) for k in _EBITDA_ADDBACKS]
            for col_idx, r in enumerate(corp_data):
                # 영업이익이 없으면 EBITDA를 계산할 수 없으므로 빈 칸으로 둔다.
                if r.get("items", {}).get(_EBITDA_OP) is None:
                    continue
                col = get_column_letter(col_idx + 2)
                terms = "+".join(f"{col}{row}" for row in (op_row, *addback_rows))
                cell = ws.cell(row=ebitda_row, column=col_idx + 2, value=f"={terms}")
                cell.number_format = "#,##0_);(#,##0);-_)"
                cell.alignment = Alignment(horizontal="right")
            ws.cell(row=ebitda_row, column=1).font = Font(bold=True)

            # 항목 컬럼 너비 조정
            ws.column_dimensions["A"].width = 14
            for col_idx in range(len(years)):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx + 2)].width = 16

            # 단위 표기
            ws.cell(row=2, column=1, value="(단위: 억원)").font = Font(
                italic=True, color="999999", size=9)

        # ── 원본 시트 (원 단위) ───────────────────────────────────────
        # 매출원가는 매출총이익 정합성 검사(파이프라인 검증)에만 쓰고 출력에서는 숨긴다.
        # EBITDA는 무형자산상각비 오른쪽에 셀 참조 수식 열로 추가한다.
        _RAW_ITEMS = [it["name"] for it in FINANCIAL_ITEMS if it["name"] != "매출원가"]
        raw_rows = []
        for r in results:
            items = r.get("items", {})

            row = {
                "기업명":       r["corp_name"],
                "연도":         r["year"],
                "corp_code":   r.get("corp_code", "-"),
                "검증":         _validation_status_label(r, use_icon=False),
                "재무제표기준": _fs_label(r),
            }
            for name in _RAW_ITEMS:
                row[name] = items.get(name)
            row["EBITDA"] = None   # 수식으로 채울 placeholder
            row["비고"] = r.get("remarks", "")
            raw_rows.append(row)
        df_raw = pd.DataFrame(raw_rows)
        df_raw.to_excel(writer, sheet_name="원본(원단위)", index=False)
        ws_raw = writer.sheets["원본(원단위)"]
        _apply_number_format(ws_raw, df_raw)

        from openpyxl.utils import get_column_letter
        # 항목 열은 F(6)부터 시작 — _RAW_ITEMS 다음에 EBITDA, 그 뒤 비고
        _FIRST_ITEM_COL = 6
        _col_of = lambda name: get_column_letter(_FIRST_ITEM_COL + _RAW_ITEMS.index(name))
        ebitda_col_idx = _FIRST_ITEM_COL + len(_RAW_ITEMS)
        ebitda_col = get_column_letter(ebitda_col_idx)
        op_col = _col_of(_EBITDA_OP)
        addback_cols = [_col_of(k) for k in _EBITDA_ADDBACKS]
        # EBITDA 수식: 영업이익 없으면 계산 불가 → 빈 칸
        for i, r in enumerate(results):
            excel_row = i + 2   # 헤더 1행 + 0-base 보정
            if r.get("items", {}).get(_EBITDA_OP) is None:
                continue
            terms = "+".join(f"{c}{excel_row}" for c in (op_col, *addback_cols))
            cell = ws_raw.cell(row=excel_row, column=ebitda_col_idx, value=f"={terms}")
            cell.number_format = "#,##0_);(#,##0);-_)"

        ws_raw.column_dimensions["A"].width = 16  # 기업명
        ws_raw.column_dimensions["B"].width = 10  # 연도
        ws_raw.column_dimensions["C"].width = 12  # corp_code
        ws_raw.column_dimensions["D"].width = 22  # 검증
        ws_raw.column_dimensions["E"].width = 14  # 재무제표기준
        # 재무 항목 + EBITDA 열
        n_item_cols = len(_RAW_ITEMS) + 1
        for i in range(n_item_cols):
            ws_raw.column_dimensions[get_column_letter(_FIRST_ITEM_COL + i)].width = 16
        # 마지막: 비고
        비고_col = get_column_letter(_FIRST_ITEM_COL + n_item_cols)
        ws_raw.column_dimensions[비고_col].width = 60
        # 검증·비고 컬럼은 줄바꿈 허용
        for col_letter in ("D", 비고_col):
            for cell in ws_raw[col_letter]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    return buf.getvalue()


# ── 검증 상세 렌더링 ──────────────────────────────────────────────────────────

def _render_validation_detail(result: dict) -> None:
    """단일 결과의 검증 상세 내역을 Streamlit으로 렌더링한다."""
    if result["status"] != "ok" or result["validation"] is None:
        st.warning(_processing_status_detail(result))
        return

    val = result["validation"]
    reason_lines = _validation_reason_lines(result)

    st.markdown("**처리 상태:**")
    st.info(_processing_status_detail(result))

    st.markdown("**검증 요약:**")
    for line in reason_lines:
        st.write(f"- {line}")

    # 검증 항목 테이블
    rows = []
    for c in val["checks"]:
        sev_icon = {"critical": "🔴", "warning": "🟡", "info": "🔵"}.get(c["severity"], "⚪")
        passed_icon = {
            "passed": "✓",
            "failed": "✗",
            "skipped_missing_data": "➖",
        }.get(c.get("status"), "✗" if c.get("passed") is False else "✓")
        exp = c.get("expected")
        act = c.get("actual")
        diff = c.get("diff")

        rows.append({
            "":       f"{passed_icon} {sev_icon}",
            "규칙":   c["rule"],
            "기대값": f"{exp:,.0f}" if isinstance(exp, (int, float)) else (str(exp) if exp else "-"),
            "실제값": f"{act:,.0f}" if isinstance(act, (int, float)) else (str(act) if act else "-"),
            "차이":   f"{diff:,.0f}원" if isinstance(diff, (int, float)) else "-",
        })

    st.dataframe(
        pd.DataFrame(rows),
        use_container_width=True,
        hide_index=True,
    )

    # 플래그
    if val["flags"]:
        st.markdown("**플래그:**")
        for flag in val["flags"]:
            st.warning(flag)
    else:
        st.success("이상 없음")

    # AI 비교 결과 (경로A·B 공통)
    ai_comp = result.get("ai_comparison")
    if ai_comp:
        source = ai_comp.get("source", "unknown")
        ai_calls = ai_comp.get("ai_calls", 0)
        source_label = {
            "agreed":          "✅ Python·AI 일치",
            "adjudicated":     "⚖️ AI 판정 (불일치 해소)",
            "ai_extract_only": "🤖 AI 추출 (판정 실패)",
            "python_fallback":  "🐍 Python만 (AI 실패)",
        }.get(source, source)
        st.markdown(f"**AI 비교 결과:** {source_label} (AI 호출: {ai_calls}회)")

        # 불일치 내역 표시
        disagreements = ai_comp.get("disagreements", {})
        if disagreements:
            comp_rows = []
            for name, (py_val, ai_val) in disagreements.items():
                final_val = result.get("items", {}).get(name)
                comp_rows.append({
                    "항목":         name,
                    "Python(원)":   f"{py_val:,.0f}" if py_val is not None else "-",
                    "AI(원)":       f"{ai_val:,.0f}" if ai_val is not None else "-",
                    "최종 채택(원)": f"{final_val:,.0f}" if final_val is not None else "-",
                })
            st.dataframe(pd.DataFrame(comp_rows), hide_index=True, use_container_width=True)

        if ai_comp.get("error"):
            st.warning(f"AI 오류: {ai_comp['error']}")

    # 재무 상세
    items = result.get("items", {})
    if any(v is not None for v in items.values()):
        st.markdown("**재무 항목 상세:**")
        detail_rows = []
        for item in FINANCIAL_ITEMS:
            val_raw = items.get(item["name"])
            detail_rows.append({
                "항목":     item["name"],
                "구분":     item["fs_type"],
                "금액(억)": _fmt_억(val_raw),
                "금액(원)": f"{val_raw:,.0f}" if val_raw is not None else "-",
            })
        ebitda_raw = _compute_ebitda(items)
        detail_rows.append({
            "항목":     "EBITDA",
            "구분":     "파생",
            "금액(억)": _fmt_억(ebitda_raw),
            "금액(원)": f"{ebitda_raw:,.0f}" if ebitda_raw is not None else "-",
        })
        st.dataframe(pd.DataFrame(detail_rows), use_container_width=True, hide_index=True)

    trace_lines = result.get("depreciation_trace") or []
    if trace_lines:
        st.markdown("**감가상각 추적 로그:**")
        st.download_button(
            label="감가상각 추적 로그 다운로드 (.txt)",
            data=_depreciation_trace_text(result),
            file_name=_depreciation_trace_filename(result),
            mime="text/plain",
            key=f"depr_trace_{result.get('corp_code', '-')}_{result.get('year', '-')}_{result.get('rcept_no', '-')}",
        )
        st.code("\n".join(trace_lines), language="text")


# ── 사이드바 ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.title("📊 DART 재무 데이터 분석기")
    st.caption("외감기업(비상장 포함) 재무제표 자동 추출 · 검증")
    st.divider()

    # ── API 키 입력 ──────────────────────────────────────────────────────
    with st.expander("🔑 API 키 설정", expanded=not get_dart_api_key()):
        st.text_input(
            "DART API Key",
            key="dart_api_key",
            help="[DART Open API](https://opendart.fss.or.kr)에서 발급받은 인증키",
        )
        st.text_input(
            "Gemini API Key (선택)",
            key="gemini_api_key",
            help="AI 교차검증용. 없으면 Python 추출만 수행",
        )

    # 입력값을 브라우저 localStorage에 저장 (다음 방문 시 자동 복원)
    _dk = json.dumps(st.session_state.get("dart_api_key", ""))
    _gk = json.dumps(st.session_state.get("gemini_api_key", ""))
    components.html(f"""
    <script>
    localStorage.setItem('dart_api_key', {_dk});
    localStorage.setItem('gemini_api_key', {_gk});
    </script>
    """, height=0)

    st.divider()

    corp_input = st.text_area(
        "기업명 입력",
        placeholder="삼성전자\n한국맥도날드\n카카오\n(줄바꿈으로 여러 기업 입력, 최대 100개)",
        height=180,
    )

    year_options = list(range(2020, 2026))  # 2020~2025
    col_y1, col_y2 = st.columns(2)
    with col_y1:
        year_start = st.selectbox("시작 연도", options=year_options, index=2)  # 기본 2022
    with col_y2:
        year_end = st.selectbox("끝 연도", options=year_options, index=5)      # 기본 2025
    years_selected = list(range(year_start, year_end + 1)) if year_start <= year_end else []

    analyze_btn = st.button("🔍 분석 시작", type="primary", use_container_width=True)


# ── 메인 영역 ─────────────────────────────────────────────────────────────────

st.header("DART 재무 데이터 분석기")

# ── 동명 법인 감지 / 분석 실행 헬퍼 ──────────────────────────────────────────

_CORP_CLS_LABEL = {"Y": "유가증권", "K": "코스닥", "N": "코넥스", "E": "기타"}


def _detect_ambiguities(corps_list: list[str]) -> dict[str, list[dict]]:
    """동일 법인명이 2개 이상인 경우 후보 정보를 모아 반환한다.

    DART corp_codes 다운로드 실패 시 예외를 그대로 전파한다(호출자가 사용자에게 표시).

    반환: {corp_name: [{corp_code, stock_code, corp_cls, ceo_nm, adres, induty_code}, ...]}
    """
    # 기업코드 캐시를 한 번에 명시적으로 로드한다 (실패 시 곧장 예외 전파).
    # 이후의 search_corp는 모두 캐시 hit이라 추가 API 호출 없이 빠르게 처리된다.
    download_corp_codes()

    info_cache: dict = st.session_state.setdefault("corp_info_cache", {})
    ambiguities: dict[str, list[dict]] = {}
    seen: set[str] = set()

    for name in corps_list:
        if name in seen or name in ambiguities:
            continue
        seen.add(name)
        df = search_corp(name)
        exact = df[df["corp_name"] == name].reset_index(drop=True)
        if len(exact) <= 1:
            continue

        codes = exact["corp_code"].tolist()
        missing = [c for c in codes if c not in info_cache]
        if missing:
            fetched = get_company_info_batch(missing)
            info_cache.update(fetched)

        rows: list[dict] = []
        for _, row in exact.iterrows():
            info = info_cache.get(row["corp_code"], {}) or {}
            rows.append({
                "corp_code":   row["corp_code"],
                "corp_name":   row["corp_name"],
                "stock_code":  row.get("stock_code") or "",
                "corp_cls":    info.get("corp_cls", ""),
                "ceo_nm":      info.get("ceo_nm", ""),
                "adres":       info.get("adres", ""),
                "induty_code": info.get("induty_code", ""),
                "est_dt":      info.get("est_dt", ""),
            })
        ambiguities[name] = rows

    return ambiguities


def _execute_analysis(
    corps_list: list[str],
    years_list: list[int],
    overrides: Optional[dict[str, str]] = None,
) -> None:
    """corps × years 곱집합에 대해 분석을 실행한다.

    overrides: {corp_name: corp_code} — 동명 법인 중 사용자가 선택한 corp_code.
    """
    overrides = overrides or {}
    tasks = list(product(corps_list, sorted(years_list)))
    total = len(tasks)

    progress_bar = st.progress(0, text="분석 준비 중...")
    status_text = st.empty()

    results: list[dict] = []
    all_logs: list[str] = []
    for i, (corp_name, year) in enumerate(tasks):
        status_text.markdown(f"**처리 중** ({i + 1}/{total}): `{corp_name}` {year}년")
        result = _analyze_one(
            corp_name, year,
            corp_code_override=overrides.get(corp_name),
        )
        all_logs.extend(result.get("analysis_log", []))
        all_logs.append("")
        results.append(result)
        progress_bar.progress((i + 1) / total, text=f"{i + 1}/{total} 완료")

    progress_bar.empty()
    status_text.empty()
    st.session_state.results = results
    st.session_state.analysis_logs = all_logs
    st.success(f"✓ {total}건 분석 완료 (성공: {sum(1 for r in results if r['status'] == 'ok')}건)")


def _clear_disambig_state() -> None:
    """동명 법인 선택 관련 세션 상태를 모두 정리한다."""
    st.session_state.pending_analysis = None
    for k in list(st.session_state.keys()):
        if k.startswith("disambig_"):
            del st.session_state[k]


# ── 분석 실행 ─────────────────────────────────────────────────────────────────
if analyze_btn:
    corps = [c.strip() for c in corp_input.splitlines() if c.strip()][:100]
    if not get_dart_api_key():
        st.warning("사이드바에서 DART API 키를 입력하거나 Streamlit secrets에 DART_API_KEY를 설정하세요.")
    elif not corps:
        st.warning("기업명을 입력하세요.")
    elif not years_selected:
        st.warning("분석 연도를 선택하세요.")
    else:
        try:
            with st.spinner("동명 법인 확인 중..."):
                ambiguities = _detect_ambiguities(corps)
        except Exception as e:
            st.error(
                "DART API에 연결할 수 없어 기업코드 목록을 가져오지 못했습니다. "
                "잠시 후 다시 시도해 주세요."
            )
            with st.expander("오류 상세"):
                st.code(f"{type(e).__name__}: {e}")
            st.stop()

        if ambiguities:
            _clear_disambig_state()
            st.session_state.pending_analysis = {
                "corps": corps,
                "years": years_selected,
                "ambiguities": ambiguities,
            }
            st.rerun()
        else:
            _clear_disambig_state()
            _execute_analysis(corps, years_selected)


# ── 동명 법인 선택 UI ────────────────────────────────────────────────────────
if st.session_state.get("pending_analysis"):
    pending = st.session_state.pending_analysis
    amb_data: dict[str, list[dict]] = pending["ambiguities"]

    st.warning(
        f"⚠️ 동일 법인명 {len(amb_data)}건이 발견되었습니다. "
        "각 회사를 정확히 지정해 주세요."
    )

    for amb_name, candidates in amb_data.items():
        st.markdown(f"#### 「{amb_name}」 — {len(candidates)}개 후보")

        labels: list[str] = []
        for c in candidates:
            stock = c.get("stock_code") or ""
            stock_disp = f"종목 {stock}" if stock else "비상장"
            cls = c.get("corp_cls") or ""
            cls_disp = _CORP_CLS_LABEL.get(cls, cls or "-")
            ceo = c.get("ceo_nm") or "-"
            adres = c.get("adres") or "-"
            labels.append(
                f"{stock_disp} [{cls_disp}] · 대표 {ceo} · {adres} "
                f"(corp_code: {c['corp_code']})"
            )

        st.radio(
            "회사 선택",
            options=list(range(len(labels))),
            format_func=lambda i, _l=labels: _l[i],
            key=f"disambig_{amb_name}",
            index=None,
        )

        with st.expander("상세 정보 표 보기"):
            st.dataframe(
                pd.DataFrame(candidates),
                use_container_width=True,
                hide_index=True,
            )

    st.divider()

    selections: dict[str, str] = {}
    for amb_name, candidates in amb_data.items():
        sel_idx = st.session_state.get(f"disambig_{amb_name}")
        if sel_idx is not None:
            selections[amb_name] = candidates[sel_idx]["corp_code"]

    all_selected = len(selections) == len(amb_data)

    col_a, col_b = st.columns(2)
    with col_a:
        if st.button(
            "✓ 선택 완료, 분석 진행",
            type="primary",
            disabled=not all_selected,
            use_container_width=True,
        ):
            corps_to_run = pending["corps"]
            years_to_run = pending["years"]
            chosen = dict(selections)
            _clear_disambig_state()
            _execute_analysis(
                corps_to_run, years_to_run,
                overrides=chosen,
            )
            st.rerun()
    with col_b:
        if st.button("취소", use_container_width=True):
            _clear_disambig_state()
            st.rerun()

# ── 결과 표시 ─────────────────────────────────────────────────────────────────
if st.session_state.results:
    results = st.session_state.results
    tab_summary, tab_detail = st.tabs(["📋 요약", "🔍 상세"])

    # ── 탭1: 요약 ─────────────────────────────────────────────────────────────
    with tab_summary:
        df_summary = _to_summary_df(results)

        # 검증 실패 행 강조
        def _highlight_row(row):
            if "✗" in str(row.get("검증", "")):
                return ["background-color: #fff0f0"] * len(row)
            if "미발견" in str(row.get("검증", "")) or "없음" in str(row.get("검증", "")):
                return ["background-color: #fffff0"] * len(row)
            return [""] * len(row)

        st.dataframe(
            df_summary.style.apply(_highlight_row, axis=1),
            use_container_width=True,
            hide_index=True,
            height=min(60 + len(df_summary) * 35, 600),
        )

        # 집계 요약
        ok_results = [r for r in results if r["status"] == "ok"]
        if ok_results:
            # 검증 통과(verified)와 추출 성공을 명확히 분리한다.
            # 스킵(필수 항목 누락)은 통과가 아니라 partial로 집계된다.
            statuses = [_result_validation_status(r) for r in ok_results]
            n_verified = statuses.count("verified")
            n_partial = statuses.count("partial")
            n_failed = statuses.count("failed_validation")

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("총 분석 건수", f"{len(results)}건")
            col2.metric("데이터 추출 성공", f"{len(ok_results)}건")
            col3.metric(
                "검증 통과", f"{n_verified}건",
                delta=(f"일부검증 {n_partial}건" if n_partial else None),
                delta_color="off",
            )
            col4.metric("검증 실패", f"{n_failed}건")

    # ── 탭2: 상세 ─────────────────────────────────────────────────────────────
    with tab_detail:
        for r in results:
            vstatus = _result_validation_status(r) if r["status"] == "ok" else None
            label_icon = {"verified": "✓", "partial": "△", "failed_validation": "✗"}.get(vstatus, "✗")
            label = f"{label_icon} {r['corp_name']} · {r['year']}년 · {r['path']} · {r['report_nm']}"
            with st.expander(label):
                _render_validation_detail(r)

    # ── 다운로드 ──────────────────────────────────────────────────────────────
    st.divider()
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        xlsx = _to_excel_bytes(results)
        st.download_button(
            label="📥 엑셀 다운로드 (요약 + 원본)",
            data=xlsx,
            file_name="dart_재무데이터.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_dl2:
        if st.session_state.get("analysis_logs"):
            from datetime import datetime
            log_text = "\n".join(st.session_state.analysis_logs)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="📋 분석 로그 다운로드 (.txt)",
                data=log_text,
                file_name=f"{ts}_dart_분석로그.txt",
                mime="text/plain",
                use_container_width=True,
            )

else:
    st.info("사이드바에서 기업명과 연도를 입력하고 '분석 시작'을 눌러주세요.")
    st.markdown("""
**사용 방법:**
1. 좌측 사이드바에 분석할 기업명을 한 줄에 하나씩 입력
2. 분석 연도 선택 (복수 선택 가능)
3. '분석 시작' 클릭

**지원 기업:**
- 상장사: 사업보고서 기반 (경로A, DART 재무제표 API)
- 비상장 외감기업: 감사보고서 기반 (경로B, DART XML 파싱)
""")
